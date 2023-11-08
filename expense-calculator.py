import tkinter as tk
from tkinter import messagebox
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# Diccionarios para almacenar los datos de ganancias y gastos
ganancias = {}
gastos = {}

# Lista de opciones de tipos de gastos
opciones_tipos = ["Alimentación", "Transporte", "Entretenimiento", "Servicios", "Otros"]
nombre_actual = ""

# Crear la ventana de registro de gasto
def crear_ventana_registro():
    ventana_registro = tk.Toplevel(ventana_principal)
    ventana_registro.title("Registrar Nuevo Gasto")
    ventana_registro.geometry("400x500")
    ventana_registro.configure(bg="#105ba6")
    
    ventana_registro_label = tk.Label(ventana_registro, text="Registrar gasto", width=400, height=3, bg="#615dc7", font=("Segoe Script", 15), relief=tk.RAISED)
    ventana_registro_label.pack(pady=20)

    # Elementos de la interfaz de usuario para el registro de gastos
    nombre_label = tk.Label(ventana_registro, text="Nombre del gasto:", bg="#615dc7", font=("Segoe Script", 10), relief=tk.RAISED)
    nombre_label.pack(pady=10)
    nombre_entry = tk.Entry(ventana_registro)
    nombre_entry.pack(pady=5)
    
    tipo_label = tk.Label(ventana_registro, text="Tipo de gasto:", bg="#615dc7", font=("Segoe Script", 10), relief=tk.RAISED)
    tipo_label.pack(pady=10)

    # Opciones de clasificación de gastos
    tipo_variable = tk.StringVar(ventana_registro)
    tipo_variable.set(opciones_tipos[0])  # Valor predeterminado
    tipo_menu = tk.OptionMenu(ventana_registro, tipo_variable, *opciones_tipos)
    tipo_menu.pack(pady=5)

    # Función para validar la entrada de texto como número
    def validate_input(new_text):
        if not new_text:
            return True
        try:
            float(new_text)
            return True
        except ValueError:
            return False

    validation = ventana_registro.register(validate_input)
    monto_label = tk.Label(ventana_registro, text="Monto del gasto:", bg="#615dc7", font=("Segoe Script", 10), relief=tk.RAISED)
    monto_label.pack(pady=10)
    monto_entry = tk.Entry(ventana_registro, validate="key", validatecommand=(validation, '%P'))
    monto_entry.pack(pady=5)

    #Funcion que controla que los entries esten completos.
    def check_empty_fields():
        if not nombre_entry.get() or not monto_entry.get():
            return False
        return True

    # Función para registrar un gasto en el diccionario de gastos y mostrar un mensaje de éxito
    def registrar_gasto():
        if not check_empty_fields():
            messagebox.showerror("Error", "Por favor complete todos los campos.")
            return

        nombre = nombre_entry.get()
        monto = float(monto_entry.get())
        tipo = tipo_variable.get()

        # Actualizar el diccionario de gastos
        if nombre in gastos:
            gastos[nombre].append((tipo, monto))
        else:
            gastos[nombre] = [(tipo, monto)]

        # Actualizar el archivo de Excel
        workbook = load_workbook("expense-calculator-data.xlsx")
        sheet = workbook["Registros"]

        # Insertar una nueva fila en la hoja de cálculo a partir de la línea 3
        sheet.insert_rows(3)

        # Agregar los datos a la nueva fila
        sheet[f"B3"] = nombre
        sheet[f"C3"] = tipo
        sheet[f"D3"] = monto

        # Guardar los cambios en el archivo
        workbook.save("expense-calculator-data.xlsx")

        messagebox.showinfo("Gasto registrado", f"Se registró un gasto de {monto} en {nombre} de tipo {tipo}.")
    
    registrar_button = tk.Button(ventana_registro, text="Registrar", command=registrar_gasto, bg="#485ebe")
    registrar_button.pack(pady=20)

    if not check_empty_fields():
        registrar_button.config(state="disabled")

    #Funcion que desactiva el boton de regsitrar si los entries no estan completos.
    def on_change(event):
        if check_empty_fields():
            registrar_button.config(state="normal")
        else:
            registrar_button.config(state="disabled")

    nombre_entry.bind("<KeyRelease>", on_change)
    monto_entry.bind("<KeyRelease>", on_change)

#Funcion para crear el grafico con los datos guardados.
def update_graph():
    workbook = load_workbook("expense-calculator-data.xlsx")
    sheet = workbook["Registros"]

    # Obtener los datos de los gastos
    gastos = {}
    for row in range(3, sheet.max_row + 1):
        nombre = sheet[f"B{row}"].value
        tipo = sheet[f"C{row}"].value
        monto = sheet[f"D{row}"].value

        if nombre in gastos:
            gastos[nombre].append((tipo, monto))
        else:
            gastos[nombre] = [(tipo, monto)]

    # Generar el gráfico
    fig, axs = plt.subplots(1, len(opciones_tipos), figsize=(6 * len(opciones_tipos), 4), squeeze=False)

    for i, tipo in enumerate(opciones_tipos):
        tipo_gastos = {}
        for nombre, detalles in gastos.items():
            for t, m in detalles:
                if t == tipo:
                    if nombre in tipo_gastos:
                        tipo_gastos[nombre] += m
                    else:
                        tipo_gastos[nombre] = m

        nombres = list(tipo_gastos.keys())
        montos = list(tipo_gastos.values())

        axs[0, i].bar(nombres, montos, color='skyblue')
        axs[0, i].set_xlabel('Nombres de Gastos')
        axs[0, i].set_ylabel('Monto total')
        axs[0, i].set_title(f'Gastos de tipo {tipo}')
        axs[0, i].tick_params(axis='x', rotation=45)

    plt.tight_layout()
    plt.show()


def ver_detalles_gasto(nombre_gasto):
    if nombre_gasto in gastos:
        detalles = gastos[nombre_gasto]
        detalles_str = "\n".join([f"Tipo: {tipo}, Monto: {monto}" for tipo, monto in detalles])
        messagebox.showinfo(f"Detalles de {nombre_gasto}", detalles_str)
    else:
        messagebox.showerror("Error", f"No se encontraron detalles para {nombre_gasto}.")

# Función para abrir la pantalla de resumen de gastos
def abrir_resumen():
    update_graph()
    
def crear_ventana_detalles():
    ventana_detalles = tk.Toplevel(ventana_principal)
    ventana_detalles.title("Detalles de Gasto")
    ventana_detalles.geometry("400x300")
    ventana_detalles.configure(bg="#105ba6")

    ventana_detalles_label = tk.Label(ventana_detalles, text="Ver Detalles de Gasto", width=400, height=3, bg="#615dc7", font=("Segoe Script", 15), relief=tk.RAISED)
    ventana_detalles_label.pack(pady=20)

    nombre_label = tk.Label(ventana_detalles, text="Nombre del gasto:", bg="#615dc7", font=("Segoe Script", 10), relief=tk.RAISED)
    nombre_label.pack(pady=10)

    nombre_entry = tk.Entry(ventana_detalles)
    nombre_entry.pack(pady=5)

    def ver_detalles():
        nombre = nombre_entry.get()
        ver_detalles_gasto(nombre)

    detalles_button = tk.Button(ventana_detalles, text="Ver Detalles", command=ver_detalles, bg="#485ebe")
    detalles_button.pack(pady=20)

def cerrar_ventana():
    ventana_principal.destroy()

#Crear la ventana principal.
ventana_principal = tk.Tk()
ventana_principal.title("Calculadora de Gastos")
ventana_principal.configure(bg="#105ba6")
ventana_principal.attributes('-fullscreen', True)

# ventana_principal.iconbitmap('icons8-calculator-16.ico')
etiqueta_bienvenida = tk.Label(ventana_principal, text="Bienvenido a la Calculadora de Gastos", width=400, height=3, bg="#615dc7", font=("Segoe Script", 15), relief=tk.RAISED)
etiqueta_bienvenida.pack(pady=20)

boton_registrar_gasto = tk.Button(ventana_principal, text="Registrar Gasto ->", compound="center", command=crear_ventana_registro, width=200, height=3, bg="#485ebe", fg="white", font=("Arial", 12), relief=tk.RAISED)
boton_ver_resumen = tk.Button(ventana_principal, text="Ver Resumen ->", compound="center", command=abrir_resumen, width=200, height=3, bg="#485ebe", fg="white", font=("Arial", 12), relief=tk.RAISED)
boton_ver_detalles = tk.Button(ventana_principal, text="Ver Detalles de Gasto ->", compound="center", command=crear_ventana_detalles, width=200, height=3, bg="#485ebe", fg="white", font=("Arial", 12), relief=tk.RAISED)
boton_cerrar_pantalla = tk.Button(ventana_principal, text="Cerrar programa", compound="center", command=cerrar_ventana, width=200, height=3, bg="#485ebe", fg="white", font=("Arial", 12), relief=tk.RAISED)

boton_registrar_gasto.pack(pady=30)
boton_ver_resumen.pack()
boton_ver_detalles.pack(pady=30)
boton_cerrar_pantalla.pack(padx=20)

ventana_principal.mainloop()
